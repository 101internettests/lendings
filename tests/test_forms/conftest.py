import os
import re
import csv
import pytest
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()


def _load_forms_urls():
    raw = os.getenv("PPROFIT_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via PPROFIT_URLS for test_forms")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in PPROFIT_URLS source")
    return urls


@pytest.fixture(params=_load_forms_urls())
def example_url(request):
    return request.param


def _load_connection_urls():
    raw = os.getenv("CONNECTION_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via CONNECTION_URLS for connection tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in CONNECTION_URLS source")
    return urls


@pytest.fixture(params=_load_connection_urls())
def connection_url(request):
    return request.param


def _load_checkaddress_urls():
    raw = os.getenv("CHECKADDRESS_BUTTON_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via CHECKADDRESS_BUTTON_URLS for checkaddress tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in CHECKADDRESS_BUTTON_URLS source")
    return urls


@pytest.fixture(params=_load_checkaddress_urls())
def checkaddress_url(request):
    return request.param

@pytest.fixture(params=_load_checkaddress_urls())
def checkaddress_button_url(request):
    return request.param

def _load_checkaddress_urls_plain():
    raw = os.getenv("CHECKADDRESS_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via CHECKADDRESS_URLS for checkaddress tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in CHECKADDRESS_URLS source")
    return urls


@pytest.fixture(params=_load_checkaddress_urls_plain())
def checkaddress_urls(request):
    return request.param

def _load_connect_cards_urls():
    raw = os.getenv("CONNECT_CARDS_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via CONNECT_CARDS_URLS for connect-cards tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in CONNECT_CARDS_URLS source")
    return urls


@pytest.fixture(params=_load_connect_cards_urls())
def connect_cards_url(request):
    return request.param


def _load_business_urls():
    raw = os.getenv("PBUSINESS_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via PBUSINESS_URLS for business tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in PBUSINESS_URLS source")
    return urls


@pytest.fixture(params=_load_business_urls())
def business_url(request):
    return request.param


def _load_business_urls_second():
    raw = os.getenv("PBUSINESS_URLS_SECOND", "").strip()
    if not raw:
        pytest.skip("No URLs provided via PBUSINESS_URLS_SECOND for business tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in PBUSINESS_URLS_SECOND source")
    return urls


@pytest.fixture(params=_load_business_urls_second())
def business_url_second(request):
    return request.param


def _load_moving_urls():
    raw = os.getenv("PMOVING_URLS", "").strip()
    urls: list[str] = []

    if raw:
        if os.path.exists(raw):
            _, ext = os.path.splitext(raw)
            ext = ext.lower()
            try:
                if ext in (".xlsx", ".xls"):
                    wb = load_workbook(raw, read_only=True, data_only=True)
                    try:
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            for cell in row:
                                if isinstance(cell, str):
                                    val = cell.strip()
                                    if val.startswith("http://") or val.startswith("https://"):
                                        urls.append(val)
                    finally:
                        wb.close()
                elif ext == ".csv":
                    with open(raw, newline="", encoding="utf-8") as f:
                        reader = csv.reader(f)
                        for row in reader:
                            for cell in row:
                                val = (cell or "").strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                else:
                    with open(raw, encoding="utf-8") as f:
                        content = f.read()
                    for token in re.split(r"[\s,]+", content):
                        val = token.strip()
                        if val and (val.startswith("http://") or val.startswith("https://")):
                            urls.append(val)
            except Exception as e:
                raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
        else:
            urls = [u.strip() for u in raw.split(",") if u.strip()]
    else:
        # Default moving URLs
        urls = [
            "https://rtk-ru.online/usluga-pereezd",
            "https://mts-home.online/usluga-pereezd",
        ]

    return urls


@pytest.fixture(params=_load_moving_urls())
def moving_url(request):
    return request.param


def _load_undecided_ttk_urls():
    raw = os.getenv("PUNDECIDED_TTK_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via PUNDECIDED_TTK_URLS for TTK undecided test")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if len(urls) < 2:
        pytest.skip("Provide at least two URLs in PUNDECIDED_TTK_URLS")
    return urls


@pytest.fixture(params=_load_undecided_ttk_urls())
def undecided_ttk_url(request):
    return request.param


def _load_undecided_urls():
    raw = os.getenv("UNDECIDED_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via UNDECIDED_URLS for undecided tests")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in UNDECIDED_URLS source")
    return urls


@pytest.fixture(params=_load_undecided_urls())
def undecided_url(request):
    return request.param



